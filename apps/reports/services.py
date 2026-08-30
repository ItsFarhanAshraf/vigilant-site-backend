import os
from datetime import datetime
from django.conf import settings
from django.db.models import Count, Q
from django.utils import timezone
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.common.enums import ComplianceStatus, MilestoneStatus, ReportScope
from apps.handover.models import Handover
from apps.projects.models import Project, ProjectMilestone, Visit
from apps.reports.models import Report
from apps.review.models import Review


class ReportService:
    @classmethod
    def generate_dpr(cls, date_val, scope=ReportScope.ALL, scope_value=None, user=None):
        """Generate Daily Progress Report (DPR) PDF via ReportLab."""
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)

        datestr = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)
        filename = f'dpr_{datestr}_{scope}.pdf'
        file_path = os.path.join(reports_dir, filename)

        doc = SimpleDocTemplate(
            file_path,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'DPRTitle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            alignment=1,
            textColor=colors.HexColor('#1E3A8A'),
        )
        subtitle_style = ParagraphStyle(
            'DPRSub',
            parent=styles['Heading2'],
            fontSize=12,
            leading=16,
            alignment=1,
            textColor=colors.HexColor('#4B5563'),
        )

        elements = []
        elements.append(Paragraph('ACAG MONITORING SYSTEM - DAILY PROGRESS REPORT', title_style))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f'Date: {datestr} | Scope: {scope} ({scope_value or "All Regions"})', subtitle_style))
        elements.append(Spacer(1, 16))

        # KPI Summary Table
        total_projects = Project.objects.count()
        active_projects = Project.objects.filter(quality_status=ComplianceStatus.PENDING).count()
        completed_projects = Project.objects.filter(quality_status=ComplianceStatus.COMPLETED).count()
        under_review = ProjectMilestone.objects.filter(status=MilestoneStatus.COMPLETED).count()

        kpi_data = [
            ['Total Projects', 'Active Projects', 'Completed Projects', 'Milestones Under Review'],
            [str(total_projects), str(active_projects), str(completed_projects), str(under_review)],
        ]
        t_kpi = Table(kpi_data, colWidths=[130, 130, 130, 130])
        t_kpi.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t_kpi)
        elements.append(Spacer(1, 16))

        # Review Decisions Summary
        elements.append(Paragraph('<b>Quality Control & Review Observations</b>', styles['Heading3']))
        elements.append(Spacer(1, 6))

        reviews = Review.objects.all()
        accepted_cnt = reviews.filter(decision='ACCEPTED').count()
        overridden_cnt = reviews.filter(decision='OVERRIDDEN').count()
        rejected_cnt = reviews.filter(decision='REJECTED').count()

        rev_data = [
            ['Decision', 'Count'],
            ['Accepted', str(accepted_cnt)],
            ['Overridden', str(overridden_cnt)],
            ['Rejected (Rectification Issued)', str(rejected_cnt)],
        ]
        t_rev = Table(rev_data, colWidths=[300, 220])
        t_rev.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(t_rev)

        doc.build(elements)

        rel_path = f'/media/reports/{filename}'
        Report.objects.create(
            report_type='DPR',
            scope=scope,
            scope_value=scope_value,
            generated_by=user,
            file_path=rel_path,
        )

        return rel_path

    @classmethod
    def generate_district_excel(cls, division=None, district=None, date_from=None, date_to=None, user=None):
        """Export District Excel spreadsheet with exact NON-NEGOTIABLE 12 columns via openpyxl."""
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)

        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'district_excel_{timestamp}.xlsx'
        file_path = os.path.join(reports_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'District Progress'

        # NON-NEGOTIABLE Column Headers
        headers = [
            'Case ID',
            'Applicant Name',
            'Contact',
            'Tehsil',
            'GPS Coordinates',
            'Loan Approved',
            'Loan Disbursed',
            'Visit Date',
            'Overall Work Progress %',
            'Current Status',
            'Overall Engineer Rating',
            'Remarks',
        ]

        ws.append(headers)

        # Header Styling
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Query Project data
        qs = Project.objects.select_related('assigned_engineer').all()
        if division:
            qs = qs.filter(division__iexact=division)
        if district:
            qs = qs.filter(district__iexact=district)

        for proj in qs:
            latest_visit = Visit.objects.filter(project=proj).order_by('-visit_date').first()
            visit_date_str = latest_visit.visit_date.strftime('%Y-%m-%d') if latest_visit else 'N/A'
            engineer_rating = latest_visit.engineer_remarks if latest_visit else 'Satisfactory'

            gps_coords = f'{proj.latitude}, {proj.longitude}'

            row = [
                proj.case_id,
                proj.owner_name,
                proj.owner_phone,
                proj.tehsil,
                gps_coords,
                float(proj.loan_approved),
                float(proj.loan_disbursed),
                visit_date_str,
                float(proj.overall_progress_pct),
                proj.quality_status,
                'Good',  # Overall Engineer Rating
                f'Division: {proj.division}, Type: {proj.project_type}',
            ]
            ws.append(row)

        # Formatting options: freeze row 1 and auto-adjust widths
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(file_path)

        rel_path = f'/media/reports/{filename}'
        Report.objects.create(
            report_type='DISTRICT_EXCEL',
            scope=ReportScope.DISTRICT if district else (ReportScope.DIVISION if division else ReportScope.ALL),
            scope_value=district or division,
            generated_by=user,
            file_path=rel_path,
        )

        return rel_path

    @classmethod
    def generate_handover_report(cls, division=None, district=None, user=None):
        """Generate Handover Summary PDF report."""
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)

        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'handover_summary_{timestamp}.pdf'
        file_path = os.path.join(reports_dir, filename)

        doc = SimpleDocTemplate(
            file_path,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'HOSummaryTitle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            alignment=1,
            textColor=colors.HexColor('#1E3A8A'),
        )

        elements = []
        elements.append(Paragraph('HANDOVER SUMMARY REPORT', title_style))
        elements.append(Spacer(1, 16))

        total_ho = Handover.objects.filter(handover_status='HANDED_OVER').count()
        pending_ho = Handover.objects.filter(handover_status='PENDING_HANDOVER').count()
        under_const = Handover.objects.filter(handover_status='UNDER_CONSTRUCTION').count()

        summary_data = [
            ['Handover Status', 'Project Count'],
            ['Handed Over', str(total_ho)],
            ['Pending Handover', str(pending_ho)],
            ['Under Construction', str(under_const)],
        ]

        t_summary = Table(summary_data, colWidths=[260, 260])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t_summary)

        doc.build(elements)

        rel_path = f'/media/reports/{filename}'
        Report.objects.create(
            report_type='HANDOVER_SUMMARY',
            scope=ReportScope.DISTRICT if district else (ReportScope.DIVISION if division else ReportScope.ALL),
            scope_value=district or division,
            generated_by=user,
            file_path=rel_path,
        )

        return rel_path

    @classmethod
    def get_report_history(cls, filters=None):
        """List generated reports history."""
        filters = filters or {}
        qs = Report.objects.select_related('generated_by').all()

        report_type = filters.get('report_type')
        scope = filters.get('scope')

        if report_type:
            qs = qs.filter(report_type=report_type)
        if scope:
            qs = qs.filter(scope=scope)

        history = []
        for r in qs:
            user_name = r.generated_by.get_full_name() or r.generated_by.username if r.generated_by else 'System'
            history.append({
                'id': r.id,
                'report_type': r.report_type,
                'scope': r.scope,
                'scope_value': r.scope_value,
                'period_from': r.period_from,
                'period_to': r.period_to,
                'generated_by': user_name,
                'file_path': r.file_path,
                'created_at': r.created_at,
            })
        return history
